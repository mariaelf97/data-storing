from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
EXCEL_FILE = "data.xlsx"

# All fields from your schema
FIELDS = [

    # General Sample Info
    "Sample_ID", "Batch", "Sample_Name", "Run_Location", "Sample_Type",

    # Sample Collection Data
    "Site_Name", "Site_Primer", "District_Name", "Epi_Week",
    "Sample_Collection_Date", "Sample_Collection_Time",
    "Metro", "Latitude", "Longitude",

    # DNA Extraction
    "Qubit_Concentration_ng_ul", "Extraction_Kit",

    # PCR
    "Mix_Name", "Master_Mix_Volume", "Forward_Primer_Volume",
    "Reverse_Primer_Volume", "Nuclease_Free_Water",
    "Denaturation_Temp", "Denaturation_Time",
    "Annealing_Temp", "Annealing_Time",
    "Extension_Temp", "Extension_Time",
    "Number_of_Cycles", "Template_DNA_Amount",

    # dPCR
    "GTP_IB_Concentration", "GTP_IB_CI_95",
    "IS6110_IB_Concentration", "IS6110_IB_CI_95",

    # Storage
    "Google_Bucket_Fastq_Location"
]


def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(FIELDS)
    wb.save(EXCEL_FILE)


@app.route("/")
def form():
    return render_template("form.html", fields=FIELDS)


@app.route("/submit", methods=["POST"])
def submit():
    if not os.path.exists(EXCEL_FILE):
        create_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    row = [request.form.get(field) for field in FIELDS]
    ws.append(row)

    wb.save(EXCEL_FILE)
    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)